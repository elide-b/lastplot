import openpyxl
from scipy import stats
import matplotlib.pyplot as plt

# Give the location of the file and the name of the sheet of interest
filepath = 'Dementia_project.xlsx'
valuesheet = ''
micesheet = ''
#TODO: mice ids -> region and mice type

wb = openpyxl.load_workbook(filepath)
ws = wb[valuesheet]
wm = wb[micesheet]

lipid_start_column = 'F'
zscore_mouse = {}
mouse_WT = {}
mean_WT = {}
std_WT = {}

z_score_sheet = wb.create_sheet('z_score_results')
for region in ws['D']:
    print(region)
    for lipid in ws['2']:
        print(lipid)
        print(ws[lipid].value)
        if ws['C'] == 'WT':
            mouse_WT[lipid].append(ws[lipid].value)
            print(mouse_WT)
        mean_WT[lipid] = mouse_WT[lipid].mean()
        std_WT[lipid] = mouse_WT[lipid].std()
        zscore_mouse[region] = (ws['C'] - mean_WT[region]) / std_WT[region]
        z_score_sheet.append([region, lipid, mouse_WT[lipid], mean_WT[lipid], std_WT[lipid], zscore_mouse[lipid]])
